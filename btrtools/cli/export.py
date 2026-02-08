"""
Data export functionality for Btrieve files.
"""

import csv
import json
import os
import sqlite3
from typing import List, Optional

from btrtools.core.btrieve import BtrieveAnalyzer, BtrieveRecord


def export_file(
    filepath: str,
    format_type: str,
    record_size: Optional[int] = None,
    max_records: Optional[int] = None,
    output_file: Optional[str] = None,
) -> str:
    """
    Export Btrieve file data to the specified format.

    Args:
        filepath: Path to the Btrieve file
        format_type: Export format ('csv', 'jsonl', or 'sqlite')
        record_size: Record size (auto-detect if None)
        max_records: Maximum records to export (None for all)
        output_file: Output file path (auto-generate if None)

    Returns:
        Path to the exported file
    """
    analyzer = BtrieveAnalyzer(filepath)

    # Auto-detect record size if not provided
    if record_size is None:
        record_size, _ = analyzer.detect_record_size()
        if record_size == 0:
            raise ValueError("Could not detect record size")

    # Extract records
    records = analyzer.extract_records(record_size, max_records)

    if not records:
        raise ValueError("No records found to export")

    # Generate output filename if not provided
    if output_file is None:
        base_name = os.path.splitext(os.path.basename(filepath))[0]
        if format_type == "csv":
            output_file = f"{base_name}.csv"
        elif format_type == "jsonl":
            output_file = f"{base_name}.jsonl"
        elif format_type == "sqlite":
            output_file = f"{base_name}.db"
        elif format_type == "excel":
            output_file = f"{base_name}.xlsx"
        elif format_type == "xml":
            output_file = f"{base_name}.xml"

    # Ensure output_file is not None
    assert output_file is not None

    # Export based on format
    if format_type == "csv":
        _export_csv(records, output_file)
    elif format_type == "jsonl":
        _export_jsonl(records, output_file)
    elif format_type == "sqlite":
        _export_sqlite(records, output_file)
    elif format_type == "excel":
        _export_excel(records, output_file)
    elif format_type == "xml":
        _export_xml(records, output_file)
    else:
        raise ValueError(f"Unsupported format: {format_type}")

    return output_file


def _export_csv(records: List[BtrieveRecord], output_file: str) -> None:
    """Export records to CSV format."""
    if not records:
        return

    # Collect all unique field names
    field_names: set[str] = set()
    for record in records:
        field_names.update(record.extracted_fields.keys())

    # Add standard fields
    standard_fields = [
        "record_num",
        "record_size",
        "decoded_text",
        "printable_chars",
        "has_digits",
        "has_alpha",
    ]
    all_fields = standard_fields + sorted(list(field_names))

    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_fields)
        writer.writeheader()

        for record in records:
            row = {
                "record_num": record.record_num,
                "record_size": record.record_size,
                "decoded_text": record.decoded_text,
                "printable_chars": record.printable_chars,
                "has_digits": record.has_digits,
                "has_alpha": record.has_alpha,
            }
            # Add extracted fields
            row.update(record.extracted_fields)
            writer.writerow(row)


def _export_jsonl(records: List[BtrieveRecord], output_file: str) -> None:
    """Export records to JSON Lines format."""
    with open(output_file, "w", encoding="utf-8") as f:
        for record in records:
            data = {
                "record_num": record.record_num,
                "record_size": record.record_size,
                "raw_bytes": record.raw_bytes.hex(),
                "decoded_text": record.decoded_text,
                "printable_chars": record.printable_chars,
                "has_digits": record.has_digits,
                "has_alpha": record.has_alpha,
                "extracted_fields": record.extracted_fields,
            }
            f.write(json.dumps(data, ensure_ascii=False) + "\n")


def _export_sqlite(records: List[BtrieveRecord], output_file: str) -> None:
    """Export records to SQLite database."""
    if not records:
        return

    # Collect all unique field names for table creation
    field_names: set[str] = set()
    for record in records:
        field_names.update(record.extracted_fields.keys())

    # Create table schema
    standard_fields = [
        ("record_num", "INTEGER"),
        ("record_size", "INTEGER"),
        ("raw_bytes", "TEXT"),
        ("decoded_text", "TEXT"),
        ("printable_chars", "INTEGER"),
        ("has_digits", "BOOLEAN"),
        ("has_alpha", "BOOLEAN"),
    ]

    # Add extracted fields (all as TEXT for flexibility)
    extracted_fields = [(name, "TEXT") for name in sorted(field_names)]

    all_fields = standard_fields + extracted_fields

    # Create table SQL
    columns_sql = ", ".join(f'"{name}" {type_}' for name, type_ in all_fields)
    create_table_sql = f"CREATE TABLE btrieve_records ({columns_sql})"

    # Insert SQL
    placeholders = ", ".join("?" for _ in all_fields)
    insert_sql = f"INSERT INTO btrieve_records VALUES ({placeholders})"  # nosec B608

    conn = sqlite3.connect(output_file)
    try:
        cursor = conn.cursor()

        # Create table
        cursor.execute(create_table_sql)

        # Insert records
        for record in records:
            values = [
                record.record_num,
                record.record_size,
                record.raw_bytes.hex(),
                record.decoded_text,
                record.printable_chars,
                record.has_digits,
                record.has_alpha,
            ]

            # Add extracted field values
            for field_name, _ in extracted_fields:
                values.append(record.extracted_fields.get(field_name, ""))

            cursor.execute(insert_sql, values)

        conn.commit()

    finally:
        conn.close()


def _export_excel(records: List[BtrieveRecord], output_file: str) -> None:
    """Export records to Excel (.xlsx) format."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    if not records:
        return

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Btrieve Records"

    # Collect all unique field names
    field_names: set[str] = set()
    for record in records:
        field_names.update(record.extracted_fields.keys())

    # Add standard fields
    standard_fields = [
        "record_num",
        "record_size",
        "raw_bytes",
        "decoded_text",
        "printable_chars",
        "has_digits",
        "has_alpha",
    ]
    all_fields = standard_fields + sorted(list(field_names))

    # Write header row with bold font
    header_font = Font(bold=True)
    for col_num, field_name in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=col_num, value=field_name)
        cell.font = header_font

    # Write data rows
    for row_num, record in enumerate(records, 2):
        # Standard fields
        ws.cell(row=row_num, column=1, value=record.record_num)
        ws.cell(row=row_num, column=2, value=record.record_size)
        ws.cell(row=row_num, column=3, value=record.raw_bytes.hex())
        ws.cell(row=row_num, column=4, value=record.decoded_text)
        ws.cell(row=row_num, column=5, value=record.printable_chars)
        ws.cell(row=row_num, column=6, value=record.has_digits)
        ws.cell(row=row_num, column=7, value=record.has_alpha)

        # Extracted fields
        col_offset = len(standard_fields)
        for i, field_name in enumerate(
            sorted(record.extracted_fields.keys()), col_offset
        ):
            ws.cell(
                row=row_num, column=i + 1, value=record.extracted_fields[field_name]
            )

    # Auto-adjust column widths
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_num).column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(output_file)


def _export_xml(records: List[BtrieveRecord], output_file: str) -> None:
    """Export records to XML format."""
    try:
        import defusedxml.minidom as minidom
        from xml.etree.ElementTree import Element, SubElement, tostring
    except ImportError:
        # xml.etree is part of standard library, so this shouldn't happen
        raise ImportError("XML support is not available")

    if not records:
        return

    # Create root element
    root = Element("btrieve_records")

    for record in records:
        record_elem = SubElement(root, "record")
        record_elem.set("number", str(record.record_num))
        record_elem.set("size", str(record.record_size))

        # Add standard fields
        SubElement(record_elem, "raw_bytes").text = record.raw_bytes.hex()
        SubElement(record_elem, "decoded_text").text = record.decoded_text
        SubElement(record_elem, "printable_chars").text = str(record.printable_chars)
        SubElement(record_elem, "has_digits").text = str(record.has_digits)
        SubElement(record_elem, "has_alpha").text = str(record.has_alpha)

        # Add extracted fields
        if record.extracted_fields:
            fields_elem = SubElement(record_elem, "extracted_fields")
            for field_name, field_value in record.extracted_fields.items():
                field_elem = SubElement(fields_elem, "field")
                field_elem.set("name", field_name)
                field_elem.text = str(field_value)

    # Pretty print XML
    rough_string = tostring(root, encoding="unicode")
    reparsed = minidom.parseString(rough_string)
    pretty_xml = reparsed.toprettyxml(indent="  ")

    # Remove extra whitespace
    lines = [line for line in pretty_xml.split("\n") if line.strip()]
    clean_xml = "\n".join(lines)

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(clean_xml)
